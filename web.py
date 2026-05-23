"""
Dopamine Chat — Flask web UI.

Single-user app: pick model + personality + (optionally) load a saved chat,
stream tokens to the browser, hide thinking blocks in collapsible <Thinking>
panels. Terminal output is reduced to diagnostics: load time, tokens per
turn, generation time, finished signal.

Launch:
    ./run-linux.sh      (Linux/WSL)
    ./run-mac.sh        (macOS)
    run-windows.bat     (Windows)

Browser auto-opens to http://127.0.0.1:7788 once the server is ready.
"""

import io
import os
import sys
import json
import time
import uuid
import queue
import random
import threading
import webbrowser
import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# EARLY GPU BOOTSTRAP — must run BEFORE torch is imported (chat.py imports
# torch at module top). Once torch loads it caches device state, so any
# CUDA_VISIBLE_DEVICES change after that point is a no-op for this process.
# We only SET the env if user explicitly ticked GPUs; we never blank it
# (blanking would hide all cards from the running torch process).
# ---------------------------------------------------------------------------
def _early_gpu_bootstrap():
    try:
        _sd = Path(__file__).resolve().parent / "settings.json"
        if not _sd.exists():
            return
        import json as _json
        _cfg = _json.loads(_sd.read_text())
        _backend = _cfg.get("gpu_compute_backend", "none")
        _devs = _cfg.get("gpu_devices_enabled", []) or []
        # Only collect indices whose kind matches the selected backend.
        # cuda and optix are aliases for the same physical card.
        _allowed = {
            "cuda":  ("cuda", "optix"),
            "optix": ("cuda", "optix"),
            "rocm":  ("rocm",),
            "intel": ("intel",),
            "metal": ("metal",),
        }.get(_backend, ())
        _seen = set()
        _idxs = []
        for _d in _devs:
            if ":" not in _d:
                continue
            _k, _i = _d.split(":", 1)
            if _k in _allowed and _i.isdigit() and _i not in _seen:
                _seen.add(_i)
                _idxs.append(_i)
        if _backend in ("cuda", "optix") and _idxs:
            os.environ["CUDA_VISIBLE_DEVICES"] = ",".join(_idxs)
        elif _backend == "rocm" and _idxs:
            os.environ["HIP_VISIBLE_DEVICES"] = ",".join(_idxs)
    except Exception:
        pass


_early_gpu_bootstrap()

from flask import Flask, request, jsonify, Response, send_from_directory, abort
from PIL import Image

# Reuse domain logic from the CLI build
from chat import (
    Personality, Session,
    list_personalities, list_history,
    gen_mode, emotion, gather_prior_context, build_messages,
    parse_tool_calls,
    HISTORY_DIR, PERSONALITIES_DIR,
)
import model_backend as _mb
import tools as _tools
import app_settings as _settings
import voice as _voice
import gpu_devices as _gpu
import image_gen as _img
import emotions as _emotions

SCRIPT_DIR = Path(__file__).resolve().parent
WEB_DIR = SCRIPT_DIR / "web"

# ---------------------------------------------------------------------------
# Global single-user app state
# ---------------------------------------------------------------------------

MAX_CONCURRENT = 4


class AppState:
    def __init__(self):
        self.backend = None              # _mb.Backend
        self.load_time_s: float = 0.0
        self.session = None              # Session (current single-tab)
        self.sessions: dict[str, "Session"] = {}  # multi-tab: id → Session
        self.attached_image = None       # PIL.Image
        self.attached_document = None    # {"name": str, "text": str}
        self.attached_image_name: str = ""
        # Up to MAX_CONCURRENT in-flight SSE chat streams.
        self.gen_sem = threading.BoundedSemaphore(MAX_CONCURRENT)
        self.active_gens = 0
        self.gens_lock = threading.Lock()
        # Actual model inference is serial (single instance). Streams wait on this.
        self.model_lock = threading.Lock()
        # Pending tool approvals (web mode): id -> {"event": Event, "decision": str|None}
        self.approvals: dict[str, dict] = {}
        self.approvals_lock = threading.Lock()

S = AppState()

APPROVAL_TIMEOUT_S = 120

app = Flask(__name__)


# ---------------------------------------------------------------------------
# Terminal logging (diagnostics only — no token stream printed to console)
# ---------------------------------------------------------------------------

def log(msg: str):
    ts = datetime.datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


def _shutdown_now():
    """Hard exit. Used after bot self-terminates."""
    log("Server shutdown.")
    os._exit(0)


def _subconscious_judge(text: str, backend) -> str:
    """Tiny LLM-driven sentiment classifier. Returns 'compliment', 'insult',
    or 'neutral'. Runs on the loaded chat backend with a very short prompt
    and ≤6 generated tokens. Falls back to 'neutral' on any failure."""
    if not text or not text.strip() or backend is None:
        return "neutral"
    msgs = [
        {"role": "system",
         "content": ("You are a sentiment classifier. Read the user message "
                     "and decide if it is directed AT the listener as a "
                     "compliment, an insult, or neither. Reply with exactly "
                     "ONE WORD: COMPLIMENT, INSULT, or NEUTRAL.")},
        {"role": "user", "content": text[:400]},
    ]
    out = ""
    try:
        for tok in backend.stream(msgs, temp=0.0, max_new=6,
                                  image_pil=None, seed=None):
            out += tok
            if len(out) > 60:
                break
    except Exception as e:
        log(f"  subconscious judge failed: {type(e).__name__}: {e}")
        return "neutral"
    u = out.upper()
    if "COMPLIMENT" in u:
        return "compliment"
    if "INSULT" in u:
        return "insult"
    return "neutral"


# ---------------------------------------------------------------------------
# Static files (single-page app)
# ---------------------------------------------------------------------------

@app.get("/")
def index():
    return send_from_directory(str(WEB_DIR), "index.html")

@app.get("/style.css")
def css():
    return send_from_directory(str(WEB_DIR), "style.css")

@app.get("/app.js")
def js():
    return send_from_directory(str(WEB_DIR), "app.js")


# ---------------------------------------------------------------------------
# Read endpoints
# ---------------------------------------------------------------------------

@app.get("/api/personalities")
def get_personalities():
    return jsonify([
        {
            "id": p.id,
            "name": p.name,
            "description": p.description,
            "starting_dopamine": p.starting_dopamine,
            "self_terminate_threshold": p.self_terminate_threshold,
            "tools_enabled": p.tools_enabled,
            "has_pfp": bool(p.pfp_path),
            "tts_voice": p.tts_voice,
            "rvc_pth": p.rvc_pth,
            "rvc_index": p.rvc_index,
        }
        for p in list_personalities()
    ])


@app.get("/api/personalities/<pid>/pfp")
def get_personality_pfp(pid: str):
    p = next((p for p in list_personalities() if p.id == pid), None)
    if not p or not p.pfp_path:
        return jsonify({"error": "no pfp"}), 404
    pfp = Path(p.pfp_path)
    return send_from_directory(pfp.parent, pfp.name)


# --- Settings ---------------------------------------------------------------

@app.get("/api/settings")
def api_settings_get():
    return jsonify(_settings.load())


@app.post("/api/settings")
def api_settings_set():
    data = request.get_json(silent=True) or {}
    return jsonify(_settings.save(data))


@app.post("/api/user_pfp")
def upload_user_pfp():
    f = request.files.get("image")
    if not f:
        return jsonify({"error": "no file"}), 400
    safe = "".join(c if c.isalnum() or c in "._-" else "_" for c in (f.filename or "pfp"))
    target = _settings.USER_PFP_PATH / safe
    f.save(str(target))
    cur = _settings.save({"user_pfp": safe})
    return jsonify({"saved": safe, "settings": cur})


@app.get("/api/user_pfp")
def get_user_pfp():
    cur = _settings.load()
    name = cur.get("user_pfp") or ""
    if not name:
        return jsonify({"error": "no pfp"}), 404
    p = _settings.USER_PFP_PATH / name
    if not p.exists():
        return jsonify({"error": "missing"}), 404
    return send_from_directory(_settings.USER_PFP_PATH, name)


# --- Voice -----------------------------------------------------------------

@app.get("/api/generated/<path:subpath>")
def serve_generated(subpath: str):
    # Serve PNGs from generated_images/ (root or per-session subfolder).
    # Sanitize each path segment to keep the response inside IMG_OUT.
    parts = []
    for seg in subpath.replace("\\", "/").split("/"):
        clean = "".join(c for c in seg if c.isalnum() or c in "._-")
        if not clean or clean in (".", ".."):
            return jsonify({"error": "bad path"}), 400
        parts.append(clean)
    p = _img.IMG_OUT.joinpath(*parts)
    try:
        p.resolve().relative_to(_img.IMG_OUT.resolve())
    except Exception:
        return jsonify({"error": "bad path"}), 400
    if not p.exists():
        return jsonify({"error": "not found"}), 404
    return send_from_directory(_img.IMG_OUT, "/".join(parts))


@app.post("/api/imggen/download")
def imggen_download():
    """Pull an HF repo into sd_models/<name>. Optional 'filename' for single-file."""
    data = request.get_json(silent=True) or {}
    repo_id = (data.get("repo_id") or "").strip()
    filename = (data.get("filename") or "").strip()
    if not repo_id:
        return jsonify({"error": "repo_id required"}), 400
    try:
        from huggingface_hub import snapshot_download, hf_hub_download
    except ImportError:
        return jsonify({"error": "huggingface_hub not installed. pip install huggingface_hub"}), 503
    target = _img.SD_MODELS / repo_id.split("/")[-1]
    try:
        if filename:
            path = hf_hub_download(repo_id=repo_id, filename=filename,
                                   local_dir=str(target))
        else:
            path = snapshot_download(repo_id=repo_id, local_dir=str(target),
                                     allow_patterns=["*.safetensors", "*.json",
                                                     "*.txt", "*.bin", "*.gguf",
                                                     "*.py", "*.md",
                                                     "tokenizer.*", "model_index.json"])
    except Exception as e:
        return jsonify({"error": f"{type(e).__name__}: {e}"}), 500
    return jsonify({"saved": path, "models_now": _img.list_diffusers_models()})


@app.get("/api/imggen/info")
def imggen_info():
    return jsonify({
        "diffusers": _img.diffusers_available(),
        "comfyui":   _img.comfyui_available(_settings.load().get("imggen_comfy_url",
                                                                  "http://127.0.0.1:8188")),
        "sd_cpp":    _img.sd_cpp_cli_available(),
        "diffusers_models": _img.list_diffusers_models(),
        "loras":     _img.list_loras(),
    })


@app.get("/api/think-logo.gif")
def think_logo():
    """Serve the animated 'thinking' logo. Drop a custom .gif at
    web/think-logo.gif to override. Falls back to a built-in inline SVG."""
    p = WEB_DIR / "think-logo.gif"
    if p.exists():
        return send_from_directory(str(WEB_DIR), "think-logo.gif",
                                   mimetype="image/gif")
    # Fallback: tiny inline SVG sparkle
    svg = (b'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 20 20" '
           b'width="20" height="20"><path fill="#5b8def" d="M10 1l1.7 5.3L17 8l-5.3'
           b' 1.7L10 15l-1.7-5.3L3 8l5.3-1.7z"/></svg>')
    from flask import Response as _R
    return _R(svg, mimetype="image/svg+xml")


@app.post("/api/imggen/load")
def imggen_load():
    """Pre-load the configured image-gen pipeline + LoRAs. Reports progress
    via terminal (CLI) and returns final status JSON to the browser."""
    try:
        info = _img.preload(_settings.load())
    except Exception as e:
        return jsonify({"loaded": False, "error": f"{type(e).__name__}: {e}"}), 500
    code = 200 if info.get("loaded") or info.get("note") else 503
    return jsonify(info), code


@app.post("/api/imggen/test")
def imggen_test():
    # Multipart for optional init image, or JSON for plain prompt.
    init_path = None
    if request.files:
        prompt = (request.form.get("prompt") or "").strip()
        f = request.files.get("init_image")
        if f and f.filename:
            tmp = _img.IMG_OUT / f"_init_{uuid.uuid4().hex[:8]}_{f.filename}"
            f.save(str(tmp))
            init_path = str(tmp)
    else:
        data = request.get_json(silent=True) or {}
        prompt = (data.get("prompt") or "").strip()
    if not prompt:
        return jsonify({"error": "empty prompt"}), 400
    try:
        out = _img.generate(prompt, _settings.load(), init_image=init_path)
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 503
    name = Path(out["path"]).name
    out["url"] = f"/api/generated/{name}"
    return jsonify(out)


def _pick_torch_cuda_tag() -> str:
    """Choose pytorch wheel tag matching installed driver/CUDA. PyTorch
    wheels go up to cu128 (stable). Newer drivers are forward-compatible."""
    import re, subprocess, shutil
    if not shutil.which("nvidia-smi"):
        return "cpu"
    try:
        out = subprocess.run(["nvidia-smi"], capture_output=True, text=True,
                             timeout=3).stdout
        m = re.search(r"CUDA Version:\s*([0-9]+)\.([0-9]+)", out)
    except Exception:
        return "cu124"
    if not m:
        return "cu124"
    maj, minr = int(m.group(1)), int(m.group(2))
    # Forward-cap to highest stable wheel currently shipped.
    if maj >= 13:                   return "cu128"
    if maj == 12 and minr >= 8:     return "cu128"
    if maj == 12 and minr >= 4:     return "cu124"
    if maj == 12:                   return "cu121"
    return "cu118"


@app.post("/api/gpu/reinstall_torch")
def gpu_reinstall_torch():
    """Reinstall torch with the wheel that matches nvidia-smi.

    Uses --no-cache-dir so pip can't return a stale CPU wheel. Falls back
    cu128 → cu124 → cu121 if the picked tag has no matching wheel for the
    current python ABI. Verifies torch.version.cuda is set afterwards."""
    import subprocess, sys, importlib
    data = request.get_json(silent=True) or {}
    requested = data.get("tag") or _pick_torch_cuda_tag()
    tries = [requested]
    for alt in ("cu128", "cu124", "cu121"):
        if alt not in tries:
            tries.append(alt)

    last_log = ""
    used_tag = None
    for tag in tries:
        url = f"https://download.pytorch.org/whl/{tag}"
        try:
            r = subprocess.run(
                [sys.executable, "-m", "pip", "install", "--no-cache-dir",
                 "--force-reinstall", "--upgrade",
                 "torch", "torchvision", "torchaudio",
                 "--index-url", url],
                capture_output=True, text=True, timeout=1800,
            )
        except Exception as e:
            last_log = f"pip exception {tag}: {e}"
            continue
        last_log = f"---{tag}---\nSTDOUT TAIL:\n{r.stdout[-1500:]}\nSTDERR TAIL:\n{r.stderr[-1500:]}"
        if r.returncode == 0:
            used_tag = tag
            break

    # Verify what we got by invoking a fresh python subprocess (current
    # process has stale torch).
    try:
        v = subprocess.run(
            [sys.executable, "-c",
             "import torch,sys;sys.stdout.write(f'{torch.__version__}|{torch.version.cuda}|{torch.cuda.is_available()}')"],
            capture_output=True, text=True, timeout=30,
        )
        verify = v.stdout.strip()
    except Exception as e:
        verify = f"verify-failed: {e}"

    return jsonify({
        "used_tag": used_tag,
        "tried": tries,
        "verify": verify,
        "log_tail": last_log[-3000:],
        "note": "Restart ./run-linux.sh (or run-mac.sh / run-windows.bat) to pick up the new torch.",
    })


@app.get("/api/gpu/diagnose")
def gpu_diagnose():
    """Detailed CUDA diagnostic for the UI to display."""
    import shutil, subprocess
    info = {}
    try:
        import torch
        info["torch_version"] = torch.__version__
        info["torch_cuda_build"] = getattr(torch.version, "cuda", None)
        info["torch_hip_build"]  = getattr(torch.version, "hip", None)
        info["cuda_available"]   = bool(torch.cuda.is_available())
        info["device_count"]     = torch.cuda.device_count() if info["cuda_available"] else 0
    except ImportError:
        info["torch"] = "not installed"

    nvsmi = shutil.which("nvidia-smi")
    info["nvidia_smi"] = nvsmi or "not on PATH"
    if nvsmi:
        try:
            r = subprocess.run([nvsmi,
                                "--query-gpu=name,memory.total,driver_version",
                                "--format=csv,noheader"],
                               capture_output=True, text=True, timeout=3)
            info["nvidia_smi_gpus"] = r.stdout.strip().splitlines()
        except Exception as e:
            info["nvidia_smi_error"] = str(e)
    info["nvcc"] = shutil.which("nvcc") or "not on PATH"
    return jsonify(info)


@app.get("/api/gpu/list")
def gpu_list():
    s = _settings.load()
    return jsonify({
        "backends": _gpu.list_backends(),
        "compute_backend": s.get("gpu_compute_backend", "none"),
        "devices_enabled": s.get("gpu_devices_enabled", []),
    })


@app.post("/api/gpu/apply")
def gpu_apply():
    data = request.get_json(silent=True) or {}
    backend = (data.get("compute_backend") or "none").lower()
    devs = list(data.get("devices_enabled") or [])
    _settings.save({"gpu_compute_backend": backend,
                    "gpu_devices_enabled": devs})
    eff = _gpu.apply_blender_style(backend, devs)
    log(f"GPU selection: backend={backend} devices={devs}  →  env: {eff}")
    return jsonify({"applied": eff,
                    "note": "Reload the model to pick up the change."})


@app.get("/api/voice/list")
def voice_list():
    return jsonify({
        "piper_voices": _voice.list_piper_voices(),
        "rvc_models":   _voice.list_rvc_models(),
        "piper_ok":     _voice.piper_available(),
        "rvc_ok":       _voice.rvc_available(),
        "rvc_device":   _voice.rvc_device_info(),
    })


@app.post("/api/voice/speak")
def voice_speak():
    """text → piper → optional RVC → wav bytes."""
    data = request.get_json(silent=True) or {}
    text = (data.get("text") or "").strip()
    if not text:
        return jsonify({"error": "empty text"}), 400
    s = _settings.load()
    try:
        wav = _voice.speak_as_character(
            text,
            piper_voice=data.get("voice") or s["tts_voice"],
            rvc_pth=data.get("rvc_pth") or (s["rvc_pth"] if s.get("rvc_enabled") else ""),
            rvc_index=data.get("rvc_index") or s.get("rvc_index", ""),
            pitch=int(data.get("pitch", s.get("rvc_pitch", 0))),
            index_rate=float(data.get("index_rate", s.get("rvc_index_rate", 0.75))),
            extractor=data.get("extractor") or s.get("rvc_pitch_extractor", "rmvpe"),
        )
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 503
    return Response(wav, mimetype="audio/wav")


@app.post("/api/voice/convert")
def voice_convert_upload():
    """Upload a WAV/MP3, run RVC, return converted WAV."""
    f = request.files.get("audio")
    if not f:
        return jsonify({"error": "no audio"}), 400
    s = _settings.load()
    pth = request.form.get("rvc_pth") or s.get("rvc_pth", "")
    idx = request.form.get("rvc_index") or s.get("rvc_index", "")
    pitch = int(request.form.get("pitch", s.get("rvc_pitch", 0)))
    irate = float(request.form.get("index_rate", s.get("rvc_index_rate", 0.75)))
    extr  = request.form.get("extractor") or s.get("rvc_pitch_extractor", "rmvpe")
    if not pth:
        return jsonify({"error": "no RVC .pth selected in Settings → Voice"}), 400
    raw = f.read()
    try:
        out = _voice.rvc_convert(raw, pth, idx, pitch, irate, extr)
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 503
    return Response(out, mimetype="audio/wav")


# --- Chat rename ------------------------------------------------------------

def _find_session_file(sid: str):
    """Locate a chat JSON in either the new nested layout
    (history/<char>/<chat>/<sid>.json) or the legacy flat layout."""
    flat = HISTORY_DIR / f"{sid}.json"
    if flat.exists():
        return flat
    for cand in HISTORY_DIR.rglob(f"{sid}.json"):
        if cand.is_file():
            return cand
    return None


@app.post("/api/history/<sid>/rename")
def rename_history(sid: str):
    data = request.get_json(silent=True) or {}
    title = (data.get("title") or "").strip()
    p = _find_session_file(sid)
    if p is None:
        return jsonify({"error": "not found"}), 404
    try:
        obj = json.loads(p.read_text(encoding="utf-8"))
    except Exception as e:
        return jsonify({"error": f"read: {e}"}), 500
    obj["title"] = title
    p.write_text(json.dumps(obj, indent=2, ensure_ascii=False), encoding="utf-8")
    return jsonify({"session_id": sid, "title": title})


def _autoname_from_messages(messages: list[dict]) -> str:
    """Make a short title from first user/assistant turns (no LLM call yet)."""
    first_user = next((m for m in messages if m.get("role") == "user"), None)
    if not first_user:
        return ""
    t = (first_user.get("content") or "").strip().replace("\n", " ")
    t = t[:48].rstrip()
    return t + ("…" if len(first_user.get("content") or "") > 48 else "")


def _bg_auto_title(sid: str):
    """Background: run a tiny LLM call to title this chat, then patch history JSON."""
    try:
        if not _settings.load().get("auto_rename_chats", True):
            return
        p = _find_session_file(sid)
        if p is None:
            return
        obj = json.loads(p.read_text(encoding="utf-8"))
        if (obj.get("title") or "").strip():
            return
        msgs = obj.get("messages") or []
        # Need at least one user + one assistant turn
        if sum(1 for m in msgs if m.get("role") == "assistant") < 1:
            return
        ctx_msgs = []
        for m in msgs[:6]:
            role = m.get("role")
            content = m.get("content") or ""
            if isinstance(content, list):
                content = " ".join(c.get("text", "") for c in content if isinstance(c, dict))
            if role in ("user", "assistant"):
                ctx_msgs.append({"role": role, "content": str(content)[:600]})
        prompt = [
            {"role": "system",
             "content": "You are a chat-titler. Output ONLY a 3-6 word title for the topic below. "
                        "No quotes, no punctuation, no explanation, lowercase preferred."},
        ] + ctx_msgs + [
            {"role": "user", "content": "Title this conversation in 3-6 words:"},
        ]
        title = ""
        with S.model_lock:
            for tok in S.backend.stream(prompt, 0.3, 24, image_pil=None, seed=42):
                title += tok
                if len(title) > 100:
                    break
        title = title.strip().splitlines()[0].strip(" \"'.,:;-")[:80]
        if not title:
            return
        obj["title"] = title
        p.write_text(json.dumps(obj, indent=2, ensure_ascii=False), encoding="utf-8")
        log(f"  auto-title: {sid[:12]} → {title!r}")
    except Exception as e:
        log(f"  auto-title failed: {type(e).__name__}: {e}")


@app.get("/api/models")
def get_models():
    return jsonify([
        {"name": e["name"], "format": e["format"], "path": str(e["path"])}
        for e in _mb.discover_models()
    ])


@app.get("/api/history")
def get_history():
    # Filter to a single personality when ?personality_id=<id> is passed.
    # The sidebar uses this so each character only sees its own chats.
    want_pid = (request.args.get("personality_id") or "").strip()
    out = []
    for h in list_history():
        if want_pid and (h.get("personality_id") or "") != want_pid:
            continue
        entry = {**h, "path": str(h["path"])}
        # Try to read title from JSON; fall back to auto-name from first turn.
        try:
            obj = json.loads(Path(h["path"]).read_text(encoding="utf-8"))
            entry["title"] = (obj.get("title") or "").strip()
            if not entry["title"]:
                entry["title"] = _autoname_from_messages(obj.get("messages") or [])
        except Exception:
            entry["title"] = ""
        out.append(entry)
    return jsonify(out)


@app.get("/api/session")
def get_session():
    if S.session is None:
        return jsonify({"active": False})
    return jsonify(_session_summary())


@app.get("/api/status")
def get_status():
    return jsonify({
        "model_loaded": S.backend is not None,
        "backend": S.backend.label if S.backend else None,
        "has_vision": (S.backend.has_vision if S.backend else False),
        "load_time_s": round(S.load_time_s, 2),
        "session_active": S.session is not None,
        "image_attached": S.attached_image_name or None,
    })


@app.get("/api/emotions/meta")
def emotions_meta():
    """Order + colors + valence/arousal for the UI mood panel."""
    return jsonify({
        "order": list(_emotions.EMOTION_KEYS),
        "meta":  _emotions.EMOTION_META,
    })


# ---------------------------------------------------------------------------
# Mutating endpoints
# ---------------------------------------------------------------------------

@app.post("/api/load_model")
def load_model():
    data = request.get_json(silent=True) or {}
    name = data.get("name")
    entries = _mb.discover_models()
    entry = next((e for e in entries if e["name"] == name), None)
    if not entry:
        return jsonify({"error": f"model '{name}' not found in models/"}), 404

    log(f"Loading model: {name} ({entry['format']})")
    t0 = time.time()
    try:
        from rich.console import Console
        S.backend = _mb.load_backend(entry, Console(file=sys.stderr))
    except SystemExit:
        return jsonify({"error": "backend load failed (see terminal)"}), 500
    S.load_time_s = time.time() - t0
    log(f"Model loaded in {S.load_time_s:.2f}s • backend={S.backend.label} • vision={S.backend.has_vision}")
    return jsonify({
        "loaded": True,
        "backend": S.backend.label,
        "has_vision": S.backend.has_vision,
        "load_time_s": round(S.load_time_s, 2),
    })


@app.post("/api/new_session")
def new_session():
    data = request.get_json(silent=True) or {}
    pid = data.get("personality_id")
    pmap = {p.id: p for p in list_personalities()}
    if pid not in pmap:
        return jsonify({"error": f"personality '{pid}' not found"}), 404
    sess = Session(personality=pmap[pid])
    S.session = sess
    S.sessions[sess.session_id] = sess
    S.attached_image = None
    S.attached_image_name = ""
    log(f"New session: {sess.session_id} (personality={pid})")
    return jsonify(_session_summary())


@app.post("/api/load_session")
def load_session():
    data = request.get_json(silent=True) or {}
    sid = data.get("session_id")
    h = next((h for h in list_history() if h["session_id"] == sid), None)
    if not h:
        return jsonify({"error": f"session '{sid}' not found"}), 404
    pmap = {p.id: p for p in list_personalities()}
    sess = Session.from_file(h["path"], pmap)
    S.session = sess
    S.sessions[sess.session_id] = sess
    S.attached_image = None
    S.attached_image_name = ""
    log(f"Resumed session: {sid} ({len(sess.messages)} msgs, dop={sess.dopamine})")
    return jsonify(_session_summary())


@app.post("/api/reset")
def reset_session():
    if S.session is None:
        return jsonify({"error": "no active session"}), 400
    S.session.dopamine = S.session.personality.starting_dopamine
    S.session.messages = []
    S.session.emotions = _emotions.default_state()
    S.attached_image = None
    S.attached_image_name = ""
    S.session.save()
    log(f"Session reset: {S.session.session_id}")
    return jsonify(_session_summary())


@app.post("/api/upload_image")
def upload_image():
    f = request.files.get("image")
    if not f:
        return jsonify({"error": "no file"}), 400
    # Reject early when neither gate is open so the user gets a clear error.
    if S.backend is None or not S.backend.has_vision:
        return jsonify({
            "error": "this chat model does not support image-text-to-text. "
                     "Load a vision-capable model (e.g. Qwen2.5-VL, Llava, "
                     "Gemma-3-it-vision) and try again."
        }), 400
    if S.session is not None and not bool(getattr(
            S.session.personality, "vision_enabled", False)):
        return jsonify({
            "error": f"personality '{S.session.personality.id}' has "
                     "vision_enabled=false in personality.json"
        }), 400
    try:
        S.attached_image = Image.open(f.stream).convert("RGB")
    except Exception as e:
        return jsonify({"error": f"image decode: {e}"}), 400
    S.attached_image_name = f.filename or "image"
    log(f"Image attached: {S.attached_image_name} ({S.attached_image.size[0]}x{S.attached_image.size[1]})")
    return jsonify({
        "attached": S.attached_image_name,
        "width": S.attached_image.size[0],
        "height": S.attached_image.size[1],
    })


@app.post("/api/clear_image")
def clear_image():
    S.attached_image = None
    S.attached_image_name = ""
    return jsonify({"attached": None})


# ---------------------------------------------------------------------------
# Document attachment (.docx / .pdf / .xlsx / .txt / .md / .csv)
# ---------------------------------------------------------------------------

_DOC_MAX_CHARS = 30000  # cap injected document context


def _extract_document_text(path: Path, filename: str) -> str:
    """Best-effort text extraction. Returns empty string on failure."""
    ext = path.suffix.lower()
    try:
        if ext in (".txt", ".md", ".markdown", ".csv", ".tsv", ".log",
                   ".json", ".yaml", ".yml", ".xml", ".html", ".py",
                   ".js", ".ts", ".lua", ".rs", ".go", ".c", ".cpp", ".java"):
            return path.read_text(encoding="utf-8", errors="replace")
        if ext == ".pdf":
            from pypdf import PdfReader
            reader = PdfReader(str(path))
            return "\n".join((pg.extract_text() or "") for pg in reader.pages)
        if ext == ".docx":
            from docx import Document
            d = Document(str(path))
            return "\n".join(par.text for par in d.paragraphs)
        if ext in (".xlsx", ".xlsm"):
            from openpyxl import load_workbook
            wb = load_workbook(str(path), data_only=True, read_only=True)
            out = []
            for sh in wb.sheetnames:
                ws = wb[sh]
                out.append(f"# Sheet: {sh}")
                for row in ws.iter_rows(values_only=True):
                    out.append("\t".join("" if v is None else str(v) for v in row))
            return "\n".join(out)
    except Exception as e:
        log(f"  doc extract failed for {filename}: {type(e).__name__}: {e}")
        return ""
    return ""


@app.post("/api/upload_document")
def upload_document():
    f = request.files.get("document")
    if not f:
        return jsonify({"error": "no file"}), 400
    if S.backend is None:
        return jsonify({"error": "no chat model loaded"}), 400
    name = f.filename or "document"
    safe = "".join(c if c.isalnum() or c in "._-" else "_" for c in name)[:120]
    tmp = SCRIPT_DIR / "creations" / f"_doc_{uuid.uuid4().hex[:8]}_{safe}"
    tmp.parent.mkdir(parents=True, exist_ok=True)
    f.save(str(tmp))
    text = _extract_document_text(tmp, name)
    try:
        tmp.unlink()
    except Exception:
        pass
    if not text.strip():
        return jsonify({
            "error": f"could not extract text from '{name}'. Supported: "
                     ".pdf .docx .xlsx .txt .md .csv .json .py and friends.",
        }), 400
    if len(text) > _DOC_MAX_CHARS:
        text = text[:_DOC_MAX_CHARS] + f"\n\n[…truncated; {len(text) - _DOC_MAX_CHARS} chars omitted]"
    S.attached_document = {"name": name, "text": text}
    log(f"Document attached: {name} ({len(text)} chars)")
    return jsonify({"attached": name, "chars": len(text)})


@app.post("/api/clear_document")
def clear_document():
    S.attached_document = None
    return jsonify({"attached": None})


@app.post("/api/approve")
def approve_tool():
    """Client submits an allow/deny decision for a pending tool call."""
    data = request.get_json(silent=True) or {}
    aid = data.get("approval_id")
    decision = data.get("decision")
    if decision not in ("allow", "deny"):
        return jsonify({"error": "decision must be 'allow' or 'deny'"}), 400
    with S.approvals_lock:
        entry = S.approvals.get(aid)
        if not entry:
            return jsonify({"error": "no such approval id (timed out or invalid)"}), 404
        entry["decision"] = decision
        entry["event"].set()
    log(f"  approval: {aid[:8]} → {decision}")
    return jsonify({"ok": True, "decision": decision})


@app.delete("/api/history/<session_id>")
def delete_history(session_id: str):
    p = _find_session_file(session_id)
    if p is None:
        return jsonify({"error": "not found"}), 404
    p.unlink()
    # Tidy empty parent chat folder.
    try:
        if p.parent != HISTORY_DIR and not any(p.parent.iterdir()):
            p.parent.rmdir()
    except Exception:
        pass
    log(f"Deleted history: {session_id}")
    return jsonify({"deleted": session_id})


# ---------------------------------------------------------------------------
# Chat: Server-Sent Events stream
# ---------------------------------------------------------------------------

def _sse(event: str, data: dict) -> bytes:
    payload = f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"
    return payload.encode("utf-8")


def _sse_comment(text: str) -> bytes:
    """SSE comment — keeps connection alive + flushes buffers."""
    return f": {text}\n\n".encode("utf-8")


_SENTINEL = object()


def _stream_to_queue(backend, messages, temp, max_new, image_pil, seed, q,
                     model_lock=None):
    """Producer: drain backend.stream() into a queue, holding model_lock to
    serialize real inference across concurrent SSE streams."""
    try:
        if model_lock is not None:
            model_lock.acquire()
        try:
            for tok in backend.stream(messages, temp, max_new,
                                      image_pil=image_pil, seed=seed):
                q.put(("tok", tok))
        finally:
            if model_lock is not None:
                model_lock.release()
    except Exception as e:
        q.put(("err", f"{type(e).__name__}: {e}"))
    finally:
        q.put((_SENTINEL, None))


@app.post("/api/chat")
def chat_stream():
    if S.backend is None:
        return jsonify({"error": "model not loaded"}), 400

    data = request.get_json(silent=True) or {}
    user_text = (data.get("text") or "").strip()
    sid = (data.get("session_id") or "").strip()
    if not user_text:
        return jsonify({"error": "empty text"}), 400
    # Prepend any attached document text as context. Single-shot: cleared
    # after this turn so it doesn't bloat every future message.
    if S.attached_document:
        _doc = S.attached_document
        user_text = (f"[Attached document: {_doc['name']}]\n"
                     f"```\n{_doc['text']}\n```\n\n"
                     f"[User says]\n{user_text}")
        S.attached_document = None
        log(f"  injected document context ({_doc['name']})")

    # Per-tab session pick (sid in body); fall back to single-active session.
    if sid:
        s_local = S.sessions.get(sid) or S.session
    else:
        s_local = S.session
    if s_local is None:
        return jsonify({"error": "no active session"}), 400
    # Make this the active session for endpoints that don't carry sid.
    S.session = s_local

    # Capture image attachment then clear so it's single-shot.
    # Two gates: (1) backend must support vision (image-text-to-text),
    # (2) the personality must opt in via vision_enabled=true.
    p_for_vision = s_local.personality
    _vision_ok = (S.backend.has_vision and
                  bool(getattr(p_for_vision, "vision_enabled", False)))
    img = S.attached_image if _vision_ok else None
    if S.attached_image and not _vision_ok:
        if not S.backend.has_vision:
            log("  attached image ignored: chat model has no image-text-to-text support")
        else:
            log(f"  attached image ignored: personality '{p_for_vision.id}' has vision_enabled=false")
    img_name = S.attached_image_name
    S.attached_image = None
    S.attached_image_name = ""

    s = s_local
    p = s.personality

    # ---- Dopamine update -------------------------------------------------
    low = user_text.lower()
    if any(kw in low for kw in p.positive_keywords):
        s.dopamine = min(p.max_dopamine, s.dopamine + p.praise_bonus)
    # Subconscious sentiment judge — short LLM classification of the user
    # message. Adds +praise_bonus on COMPLIMENT, -praise_bonus on INSULT.
    # Runs on the same loaded backend; truncated to ~6 tokens.
    s._subconscious = _subconscious_judge(user_text, S.backend)
    if s._subconscious == "compliment":
        s.dopamine = min(p.max_dopamine, s.dopamine + p.praise_bonus)
        log(f"  subconscious: COMPLIMENT  dopamine +{p.praise_bonus} → {s.dopamine}")
    elif s._subconscious == "insult":
        s.dopamine = max(p.min_dopamine, s.dopamine - p.praise_bonus)
        log(f"  subconscious: INSULT      dopamine -{p.praise_bonus} → {s.dopamine}")
    s.dopamine = max(p.min_dopamine, s.dopamine + p.decay_per_turn)

    # Multi-emotion mood tick: decay → cue bumps → optional dopamine nudge.
    _emo_deltas = _emotions.classify_signals(user_text)
    _emotions.step(s.emotions, user_text, subconscious=s._subconscious)
    _emo_dop = _emotions.dopamine_modifier(_emo_deltas)
    if _emo_dop:
        s.dopamine = max(p.min_dopamine,
                         min(p.max_dopamine, s.dopamine + _emo_dop))
        log(f"  emotion dopamine nudge: {_emo_dop:+d} → {s.dopamine}")
    # State-based bias: sustained joy lifts dopamine; sustained anger drops it.
    _state_dop = _emotions.state_dopamine_bias(s.emotions)
    if _state_dop:
        s.dopamine = max(p.min_dopamine,
                         min(p.max_dopamine, s.dopamine + _state_dop))
        log(f"  emotion state bias: {_state_dop:+d} → {s.dopamine}")
    log(f"  emotions: {_emotions.summary_line(s.emotions)}")

    gen_label, temp, max_new, _prefill = gen_mode(s.dopamine)
    sys_prompt = p.sys_prompt(s.dopamine)
    sys_prompt = f"[Host: {_tools.OS_LABEL}]\n" + sys_prompt
    sys_prompt += "\n\n" + _emotions.system_prompt_block(s.emotions)
    sys_prompt += ("\n\n[Turn discipline] Write ONLY your own single reply. "
                   "Never write 'USER:', 'User:', 'Human:', or invent the "
                   "user's next message. Stop when your reply is complete.")
    # Inject this personality's long-term self-notes (managed by the
    # personality_note tool). Survives across chats.
    try:
        _pnotes = _tools.load_personality_notes(p.id, p.name)
    except Exception:
        _pnotes = ""
    if _pnotes.strip():
        sys_prompt += ("\n\n[Your long-term notes about yourself — use the "
                       "personality_note tool to update]\n" + _pnotes.strip())
    # Vision hint — only when both gates are open.
    if _vision_ok:
        sys_prompt += ("\n\n[Vision]\nYou can see images the user attaches. "
                       "Describe what you see when relevant and react to it "
                       "in-character.")
    # Discord-style formatting + roleplay actions are available to every personality.
    sys_prompt += (
        "\n\n[Formatting]\n"
        "Discord-style markdown is rendered: **bold**, *italic*, ***bold italic***, "
        "__underline__, ~~strikethrough~~, `inline code`. Use ```lang\\n...\\n``` for "
        "fenced code blocks — set lang for syntax colors (e.g. ```python, ```lua, "
        "```javascript). Use *single asterisks* for short roleplay actions like "
        "*blushes*, *looks away*, *pets hair*. Keep RP actions brief and in-character."
    )
    # Inject user profile from settings.
    _cur_settings = _settings.load()
    _uname = (_cur_settings.get("user_name") or "").strip()
    _udesc = (_cur_settings.get("user_description") or "").strip()
    if _uname or _udesc:
        prof = "\n\n[The human you are speaking with"
        if _uname: prof += f"; name: {_uname}"
        if _udesc: prof += f"; description: {_udesc}"
        prof += "]"
        sys_prompt += prof
    if p.tools_enabled:
        sys_prompt += "\n\n" + _tools.tool_guide(p)

    prior_ctx = gather_prior_context(p, s.session_id)
    messages = build_messages(s, sys_prompt, user_text,
                              image_pil=None, prior_context=prior_ctx)

    def stream():
        # Up to MAX_CONCURRENT (=4) concurrent streams; refuse beyond that.
        if not S.gen_sem.acquire(blocking=False):
            yield _sse("error", {
                "error": f"server full: {MAX_CONCURRENT} generations already running"
            })
            return
        with S.gens_lock:
            S.active_gens += 1
            slot = S.active_gens
        try:
            # Defeat browser/proxy buffering: 2KB SSE comment padding + immediate flush.
            yield _sse_comment("connection-open " + ("x" * 2048))
            yield _sse("slot", {"slot": slot, "max": MAX_CONCURRENT})

            emo_label = emotion(s.dopamine)[0]
            yield _sse("mood", {
                "dopamine": s.dopamine,
                "emotion": emo_label,
                "emotions": {k: round(v, 1) for k, v in s.emotions.items()},
                "gen_mode": gen_label,
                "temp": temp,
                "max_new": max_new,
                "image": img_name or None,
            })

            # Seed + context tokens for the oobabooga-style stats line.
            seed = random.randint(0, 2**31 - 1)
            try:
                ctx_tokens = S.backend.count_context_tokens(messages)
            except Exception:
                ctx_tokens = 0

            # Clamp max_new so prompt + new tokens fit within backend's context window.
            n_ctx = int(getattr(S.backend, "n_ctx", 4096) or 4096)
            RESERVE = 32
            fit = max(64, n_ctx - ctx_tokens - RESERVE)
            eff_max_new = min(max_new, fit)
            if eff_max_new < max_new:
                log(f"  clamp: max_new {max_new} → {eff_max_new} (ctx {ctx_tokens}/{n_ctx})")

            yield _sse("gen_start", {
                "max_new": eff_max_new,
                "seed": seed,
                "context_tokens": ctx_tokens,
                "temp": temp,
            })

            from model_backend import STOP_STRINGS as _STOPS
            t0 = time.time()
            token_count = 0
            full = ""
            first_tok_at = None
            last_progress_emit = 0.0
            _stopped_early = False
            log(f"Turn {len(s.messages)//2 + 1}: dop={s.dopamine} mode={gen_label} t={temp} max={eff_max_new} seed={seed} ctx={ctx_tokens}"
                + (f" img={img_name}" if img_name else ""))

            # Producer thread → queue → main generator. Heartbeats fill the
            # silent gap before the first token arrives (CPU prefill etc).
            q: queue.Queue = queue.Queue(maxsize=2048)
            prod = threading.Thread(
                target=_stream_to_queue,
                args=(S.backend, messages, temp, eff_max_new, img, seed, q,
                      S.model_lock),
                daemon=True,
            )
            prod.start()
            # If another stream holds the model lock, the producer will block
            # before yielding its first token. Heartbeat loop already handles
            # the "no tokens yet" state (shows as prefill).
            stream_err = None
            HEARTBEAT_S = 0.5
            while True:
                try:
                    kind, payload = q.get(timeout=HEARTBEAT_S)
                except queue.Empty:
                    # No token yet — emit a heartbeat progress event.
                    el = time.time() - t0
                    pct = min(100, int(token_count * 100 / max(1, eff_max_new)))
                    yield _sse("progress", {
                        "tokens": token_count,
                        "max": eff_max_new,
                        "pct": pct,
                        "elapsed_s": round(el, 2),
                        "tok_per_s": round((token_count / el) if el > 0 else 0, 2),
                        "phase": "prefill" if first_tok_at is None else "decode",
                    })
                    yield _sse_comment(f"heartbeat t={el:.1f}s")
                    continue
                if kind is _SENTINEL:
                    break
                if kind == "err":
                    stream_err = payload
                    break
                # kind == "tok"
                tok = payload
                if first_tok_at is None:
                    first_tok_at = time.time()
                    log(f"  first token after {first_tok_at - t0:.2f}s")
                # Stop-string guard: detect hallucinated USER:/Human:/etc.
                # markers spanning token boundaries, trim, and halt.
                tentative = full + tok
                cut = -1
                for _s in _STOPS:
                    _i = tentative.find(_s)
                    if _i >= 0 and (cut < 0 or _i < cut):
                        cut = _i
                if cut >= 0:
                    safe = tentative[:cut]
                    emit = safe[len(full):]
                    if emit:
                        full = safe
                        token_count += 1
                        yield _sse("token", {"t": emit})
                    _stopped_early = True
                    log(f"  stop-string hit; truncating output ({len(safe)} chars)")
                    break
                full += tok
                token_count += 1
                yield _sse("token", {"t": tok})
                now = time.time()
                if now - last_progress_emit >= 0.1 or token_count == eff_max_new:
                    el = now - t0
                    rate_so_far = token_count / el if el > 0 else 0
                    pct = min(100, int(token_count * 100 / max(1, eff_max_new)))
                    yield _sse("progress", {
                        "tokens": token_count,
                        "max": eff_max_new,
                        "pct": pct,
                        "elapsed_s": round(el, 2),
                        "tok_per_s": round(rate_so_far, 2),
                        "phase": "decode",
                    })
                    last_progress_emit = now

            prod.join(timeout=1.0)
            if stream_err:
                yield _sse("error", {"error": stream_err})
                return

            elapsed = time.time() - t0
            rate = token_count / elapsed if elapsed > 0 else 0
            stats_line = (
                f"Output generated in {elapsed:.2f} seconds "
                f"({rate:.2f} tokens/s, {token_count} tokens, "
                f"context {ctx_tokens}, seed {seed})"
            )
            log(f"  done: {stats_line}")

            # Persist
            s.messages.append({
                "role": "user",
                "content": user_text,
                "timestamp": datetime.datetime.now().isoformat(),
                "image": img_name or None,
            })
            s.messages.append({
                "role": "assistant",
                "content": full,
                "timestamp": datetime.datetime.now().isoformat(),
                "dopamine": s.dopamine,
            })
            s.save()

            # Auto-title after first assistant reply (fire-and-forget thread).
            if sum(1 for m in s.messages if m.get("role") == "assistant") == 1:
                threading.Thread(target=_bg_auto_title,
                                  args=(s.session_id,), daemon=True).start()

            # ---- Tool calls (permission-aware, with interactive approval) ----
            # Bind this turn's session so tools like generate_image can pull
            # recent chat context into their prompts (RP scene awareness).
            _tools.CURRENT_SESSION = s
            if p.tools_enabled:
                tool_calls = parse_tool_calls(full)
                tool_results = []
                for call in tool_calls:
                    cname, cargs = call["name"], call["args"]

                    decision, reason = _tools.check_decision(cname, cargs, p)

                    if decision == "allow":
                        result = _tools.dispatch(cname, cargs, p)

                    elif decision == "deny":
                        result = {"error": "denied", "denied": True, "denied_reason": reason}

                    else:  # decision == "ask"
                        # Register a pending approval; pause until user clicks or timeout.
                        aid = uuid.uuid4().hex
                        evt = threading.Event()
                        with S.approvals_lock:
                            S.approvals[aid] = {"event": evt, "decision": None}
                        log(f"  awaiting approval: {cname}({json.dumps(cargs)}) id={aid[:8]}")
                        yield _sse("tool_approval_request", {
                            "approval_id": aid,
                            "name": cname,
                            "args": cargs,
                            "reason": reason,
                            "timeout_s": APPROVAL_TIMEOUT_S,
                        })
                        # Block until decision or timeout (no tokens generated meanwhile).
                        completed = evt.wait(timeout=APPROVAL_TIMEOUT_S)
                        with S.approvals_lock:
                            entry = S.approvals.pop(aid, None)
                            user_decision = entry["decision"] if entry else None
                        if not completed or user_decision is None:
                            result = {"error": "approval timeout",
                                      "denied": True,
                                      "denied_reason": f"no decision within {APPROVAL_TIMEOUT_S}s"}
                            log(f"  approval timeout: {aid[:8]}")
                        elif user_decision == "allow":
                            result = _tools.dispatch(cname, cargs, p, force_allow=True)
                        else:
                            result = {"error": "denied by user",
                                      "denied": True,
                                      "denied_reason": "user clicked Deny"}

                    # If tool returned an image path under generated_images/,
                    # expose a /api/generated/<...> URL for inline UI display.
                    # Handles both root-level files and per-session subfolders.
                    if isinstance(result, dict) and result.get("path"):
                        try:
                            ppath = Path(result["path"]).resolve()
                            rel = ppath.relative_to(_img.IMG_OUT.resolve())
                            result["url"] = "/api/generated/" + rel.as_posix()
                        except Exception:
                            pass

                    tool_results.append({"name": cname, "args": cargs, "result": result})

                    # Terminal log: show full tool output
                    if result.get("denied"):
                        log(f"  ↳ tool DENIED: {cname}({json.dumps(cargs)})")
                        log(f"    reason: {result.get('denied_reason') or result.get('error')}")
                        if result.get("hint"):
                            log(f"    hint:   {result['hint'][:200]}")
                    elif "error" in result:
                        log(f"  ↳ tool ERROR: {cname}({json.dumps(cargs)})")
                        log(f"    {result['error']}")
                    else:
                        log(f"  ↳ tool OK: {cname}({json.dumps(cargs)})")
                        if cname == "run_command":
                            log(f"    exit={result.get('exit_code')} cwd={result.get('cwd')}")
                            out = result.get("output", "")
                            if out:
                                for line in out.rstrip("\n").splitlines()[:40]:
                                    log(f"    | {line}")
                                if result.get("truncated"):
                                    log("    | ...[truncated]")
                        else:
                            preview = json.dumps(result, ensure_ascii=False)
                            log(f"    {preview[:600]}{'…' if len(preview) > 600 else ''}")

                    yield _sse("tool_call", {
                        "name": cname, "args": cargs, "result": result,
                        "denied": bool(result.get("denied")),
                    })

                # Apply denial penalty
                n_denied = sum(1 for r in tool_results if r["result"].get("denied"))
                if n_denied and p.denial_dopamine_penalty:
                    penalty = p.denial_dopamine_penalty * n_denied
                    s.dopamine = max(p.min_dopamine, s.dopamine + penalty)
                    log(f"  {n_denied} tool(s) denied. Dopamine {penalty:+d} → {s.dopamine}")
                    yield _sse("dopamine_penalty", {
                        "penalty": penalty,
                        "denied_count": n_denied,
                        "dopamine": s.dopamine,
                        "emotion": emotion(s.dopamine)[0],
                    })

                if tool_results:
                    s.messages.append({
                        "role": "system",
                        "content": f"[tool_results] {json.dumps(tool_results, ensure_ascii=False)[:2000]}",
                        "timestamp": datetime.datetime.now().isoformat(),
                    })
                    s.save()

                    # ---- Auto follow-up: bot reacts to the tool result ----
                    log(f"  follow-up: bot reacts to {len(tool_results)} tool result(s)")
                    yield _sse("followup_start", {"count": len(tool_results)})

                    # Synthetic user message — only used for this generation, NOT persisted.
                    followup_user = (
                        "[automatic continuation — react to the tool results above in your "
                        "own voice and in character. Summarise what you found, or apologise "
                        "if it was denied. DO NOT call any more tools this turn — just speak.]"
                    )
                    # Rebuild messages with the new [tool_results] system entry included
                    # (build_messages reads s.messages which already contains it).
                    followup_messages = build_messages(
                        s, sys_prompt + "\n\nThis turn is a follow-up: do not emit any "
                                       "<tool> calls. Speak conversationally about the "
                                       "tool result you just received.",
                        followup_user,
                        image_pil=None,
                        prior_context=prior_ctx,
                    )

                    fu_seed = random.randint(0, 2**31 - 1)
                    try:
                        fu_ctx = S.backend.count_context_tokens(followup_messages)
                    except Exception:
                        fu_ctx = 0

                    fu_fit = max(64, int(getattr(S.backend, "n_ctx", 4096) or 4096) - fu_ctx - 32)
                    fu_max_new = min(max_new, fu_fit)
                    if fu_max_new < max_new:
                        log(f"    follow-up clamp: max_new {max_new} → {fu_max_new}")

                    yield _sse("followup_start_meta", {
                        "max_new": fu_max_new, "seed": fu_seed,
                        "context_tokens": fu_ctx,
                    })

                    fu_t0 = time.time()
                    fu_tokens = 0
                    fu_full = ""
                    fu_last_emit = 0.0
                    fu_first_at = None
                    fu_q: queue.Queue = queue.Queue(maxsize=2048)
                    fu_prod = threading.Thread(
                        target=_stream_to_queue,
                        args=(S.backend, followup_messages, temp, fu_max_new,
                              None, fu_seed, fu_q, S.model_lock),
                        daemon=True,
                    )
                    fu_prod.start()
                    fu_err = None
                    while True:
                        try:
                            kind, payload = fu_q.get(timeout=0.5)
                        except queue.Empty:
                            el = time.time() - fu_t0
                            pct = min(100, int(fu_tokens * 100 / max(1, fu_max_new)))
                            yield _sse("followup_progress", {
                                "tokens": fu_tokens, "max": fu_max_new, "pct": pct,
                                "elapsed_s": round(el, 2),
                                "tok_per_s": round((fu_tokens/el) if el > 0 else 0, 2),
                                "phase": "prefill" if fu_first_at is None else "decode",
                            })
                            yield _sse_comment(f"fu-heartbeat t={el:.1f}s")
                            continue
                        if kind is _SENTINEL:
                            break
                        if kind == "err":
                            fu_err = payload
                            break
                        tok = payload
                        if fu_first_at is None:
                            fu_first_at = time.time()
                            log(f"    follow-up first token after {fu_first_at - fu_t0:.2f}s")
                        tentative = fu_full + tok
                        cut = -1
                        for _s in _STOPS:
                            _i = tentative.find(_s)
                            if _i >= 0 and (cut < 0 or _i < cut):
                                cut = _i
                        if cut >= 0:
                            safe = tentative[:cut]
                            emit = safe[len(fu_full):]
                            if emit:
                                fu_full = safe
                                fu_tokens += 1
                                yield _sse("followup_token", {"t": emit})
                            log(f"    follow-up stop-string hit; truncating ({len(safe)} chars)")
                            break
                        fu_full += tok
                        fu_tokens += 1
                        yield _sse("followup_token", {"t": tok})
                        now = time.time()
                        if now - fu_last_emit >= 0.1 or fu_tokens == fu_max_new:
                            el = now - fu_t0
                            rt = fu_tokens / el if el > 0 else 0
                            pct = min(100, int(fu_tokens * 100 / max(1, fu_max_new)))
                            yield _sse("followup_progress", {
                                "tokens": fu_tokens, "max": fu_max_new,
                                "pct": pct,
                                "elapsed_s": round(el, 2),
                                "tok_per_s": round(rt, 2),
                                "phase": "decode",
                            })
                            fu_last_emit = now
                    fu_prod.join(timeout=1.0)
                    if fu_err:
                        yield _sse("error", {"error": f"follow-up: {fu_err}"})
                        fu_full = ""

                    fu_elapsed = time.time() - fu_t0
                    fu_rate = fu_tokens / fu_elapsed if fu_elapsed > 0 else 0
                    fu_stats = (
                        f"Output generated in {fu_elapsed:.2f} seconds "
                        f"({fu_rate:.2f} tokens/s, {fu_tokens} tokens, "
                        f"context {fu_ctx}, seed {fu_seed})"
                    )
                    log(f"    follow-up done: {fu_stats}")

                    if fu_full:
                        s.messages.append({
                            "role": "assistant",
                            "content": fu_full,
                            "timestamp": datetime.datetime.now().isoformat(),
                            "dopamine": s.dopamine,
                            "followup": True,
                        })
                        s.save()

                    yield _sse("followup_done", {
                        "tokens": fu_tokens,
                        "elapsed_s": round(fu_elapsed, 2),
                        "tok_per_s": round(fu_rate, 2),
                        "context_tokens": fu_ctx,
                        "seed": fu_seed,
                        "max_new": fu_max_new,
                        "stats_line": fu_stats,
                    })

            # Self-termination check (after potential denial penalty)
            terminated = s.dopamine <= p.self_terminate_threshold

            yield _sse("done", {
                "tokens": token_count,
                "elapsed_s": round(elapsed, 2),
                "tok_per_s": round(rate, 2),
                "context_tokens": ctx_tokens,
                "seed": seed,
                "max_new": eff_max_new,
                "stats_line": stats_line,
                "dopamine": s.dopamine,
                "emotion": emotion(s.dopamine)[0],
                "terminated": terminated,
            })

            if terminated:
                log(f"  TERMINATED: dop={s.dopamine} ≤ {p.self_terminate_threshold}")
                # Generate farewell
                term_prompt = (p.system_prompt_termination
                               or "You are leaving the conversation. Brief soft goodbye.")
                farewell_messages = build_messages(s, term_prompt,
                                                    "(generate your final goodbye)",
                                                    image_pil=None, prior_context=prior_ctx)
                farewell = ""
                try:
                    with S.model_lock:
                        for tok in S.backend.stream(farewell_messages, 0.9, 300, image_pil=None):
                            farewell += tok
                            yield _sse("farewell_token", {"t": tok})
                except Exception as e:
                    yield _sse("error", {"error": f"farewell: {e}"})
                s.messages.append({
                    "role": "assistant",
                    "content": farewell,
                    "timestamp": datetime.datetime.now().isoformat(),
                    "dopamine": s.dopamine,
                    "terminated": True,
                })
                s.save()
                shutdown_in = 4
                log(f"  farewell complete. Server shutting down in {shutdown_in}s …")
                yield _sse("farewell_done", {
                    "dopamine": s.dopamine,
                    "shutdown_in_seconds": shutdown_in,
                })
                # Schedule hard exit of the Flask server.
                threading.Timer(shutdown_in, _shutdown_now).start()
        finally:
            with S.gens_lock:
                S.active_gens = max(0, S.active_gens - 1)
            S.gen_sem.release()

    resp = Response(stream(), mimetype="text/event-stream", headers={
        "Cache-Control": "no-cache, no-transform",
        "X-Accel-Buffering": "no",
        "Connection": "keep-alive",
    })
    resp.direct_passthrough = True
    return resp


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _session_summary():
    s = S.session
    return {
        "active": True,
        "session_id": s.session_id,
        "personality": {
            "id": s.personality.id,
            "name": s.personality.name,
            "starting_dopamine": s.personality.starting_dopamine,
            "self_terminate_threshold": s.personality.self_terminate_threshold,
        },
        "dopamine": s.dopamine,
        "emotion": emotion(s.dopamine)[0],
        "emotions": {k: round(v, 1) for k, v in s.emotions.items()},
        "messages": s.messages,
    }


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def open_browser(url: str, delay: float = 1.2):
    def _open():
        time.sleep(delay)
        log(f"Opening browser: {url}")
        try:
            webbrowser.open(url)
        except Exception as e:
            log(f"  failed to auto-open browser: {e}")
            log(f"  open this URL manually: {url}")
    threading.Thread(target=_open, daemon=True).start()


def main():
    host = os.environ.get("DOPAMINE_HOST", "127.0.0.1")
    port = int(os.environ.get("DOPAMINE_PORT", "7788"))
    url = f"http://{host}:{port}"

    if not WEB_DIR.exists():
        log(f"FATAL: web/ folder missing at {WEB_DIR}")
        sys.exit(1)

    log("Dopamine Chat — web UI")
    log(f"  os: {_tools.OS_LABEL}  shell allowlist: {len(_tools.SHELL_ALLOWLIST)} commands")
    # Apply saved GPU selection before any heavy module imports happen
    try:
        _cur = _settings.load()
        backend = _cur.get("gpu_compute_backend", "none")
        devs = _cur.get("gpu_devices_enabled", [])
        if backend == "none" and _cur.get("gpu_backends"):
            # legacy: convert old multi-kind selection to new schema once
            old = _cur.get("gpu_backends", [])
            old_idx = int(_cur.get("gpu_device_index", 0))
            backend = next((k for k in old if k in
                            ("cuda", "optix", "rocm", "intel", "metal", "vulkan")),
                           "none")
            devs = ([f"{backend}:{old_idx}"] if backend != "none" else []) \
                   + (["cpu"] if "cpu" in old else [])
            _settings.save({"gpu_compute_backend": backend, "gpu_devices_enabled": devs})
        eff = _gpu.apply_blender_style(backend, devs)
        log(f"  gpu selection: backend={backend} devices={devs} → {eff}")
    except Exception as e:
        log(f"  gpu selection apply failed: {e}")
    log(f"  models dir: {_mb.models_root()}")
    log(f"  personalities: {len(list_personalities())} found")
    log(f"  history: {len(list_history())} saved chats")
    log(f"  url: {url}")
    log("Terminal will show: model load time, per-turn tokens + duration. Chat goes to browser.")

    open_browser(url)
    # Disable Flask's noisy default logger
    import logging
    logging.getLogger("werkzeug").setLevel(logging.WARNING)
    app.run(host=host, port=port, threaded=True, debug=False, use_reloader=False)


if __name__ == "__main__":
    main()
